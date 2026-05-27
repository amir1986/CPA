"""Excel (.xlsx) extractor for Trial Balance and General Ledger.

Header detection is heuristic — it looks for a row that contains
canonical column names (case-insensitive substring match). Numbers
are parsed with thousands separators tolerated.

Returns dictionaries ready to upsert into ``trial_balances`` or
``gl_entries`` (see ``app/ingest_docs/canonicalize.py``).
"""

from __future__ import annotations

import datetime as dt
import io
import re
from dataclasses import dataclass
from typing import Any

TB_HEADERS = {
    "account_code": ["code", "account code", "acct", "acct code", "account no"],
    "account_name": ["account", "name", "account name", "description"],
    "opening": ["opening", "open balance", "beg balance", "beginning"],
    "debit_total": ["debit", "dr", "debits"],
    "credit_total": ["credit", "cr", "credits"],
    "closing": ["closing", "close balance", "ending", "end balance", "balance"],
}

GL_HEADERS = {
    "je_number": ["je", "journal", "entry no", "voucher", "je number", "doc no"],
    "je_date": ["date", "journal date", "trans date", "transaction date"],
    "posting_date": ["posting date", "posted"],
    "account_code": ["account", "code", "acct", "account code"],
    "description": ["description", "memo", "narrative", "detail"],
    "debit": ["debit", "dr"],
    "credit": ["credit", "cr"],
    "amount": ["amount", "net"],
    "preparer": ["preparer", "user", "entered by"],
    "approver": ["approver", "approved by"],
    "currency": ["currency", "ccy", "cur"],
}


@dataclass(frozen=True)
class TBRow:
    account_code: str | None
    account_name: str | None
    opening: float
    debit_total: float
    credit_total: float
    closing: float


@dataclass(frozen=True)
class GLRow:
    je_number: str | None
    je_date: dt.date | None
    posting_date: dt.date | None
    account_code: str | None
    description: str | None
    debit: float
    credit: float
    currency: str | None
    preparer: str | None
    approver: str | None


def parse_amount(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    if s == "" or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(v: Any) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _header_index(header_row: list[str], canonical: dict[str, list[str]]) -> dict[str, int]:
    """Map canonical → column-index by matching aliases against header cells.

    Priority (per key):
      1. Cell equals an alias exactly (case-insensitive).
      2. Cell contains an alias as a whole word (no inside-word matches —
         "cr" must not match inside "description").
    """
    mapping: dict[str, int] = {}
    lowered = [re.sub(r"\s+", " ", str(h or "").strip().lower()) for h in header_row]
    cell_words: list[set[str]] = [set(re.findall(r"[\w']+", c)) for c in lowered]
    for key, aliases in canonical.items():
        idx = _match_alias(lowered, cell_words, aliases)
        if idx is not None:
            mapping[key] = idx
    return mapping


def _match_alias(lowered: list[str], cell_words: list[set[str]], aliases: list[str]) -> int | None:
    aliases_lc = [a.lower().strip() for a in aliases]
    for idx, cell in enumerate(lowered):
        if cell in aliases_lc:
            return idx
    for idx, words in enumerate(cell_words):
        for alias in aliases_lc:
            atoks = alias.split()
            if len(atoks) > 1:
                if all(tok in words for tok in atoks):
                    return idx
            elif alias in words:
                return idx
    return None


def _detect_header_row(rows: list[list[Any]], canonical: dict[str, list[str]], *, scan: int = 15) -> int | None:
    for i, row in enumerate(rows[:scan]):
        idx = _header_index([str(c or "") for c in row], canonical)
        if len(idx) >= 3:
            return i
    return None


def extract_trial_balance(xlsx_bytes: bytes) -> list[TBRow]:
    rows = _read_first_sheet(xlsx_bytes)
    header_at = _detect_header_row(rows, TB_HEADERS)
    if header_at is None:
        return []
    cols = _header_index([str(c or "") for c in rows[header_at]], TB_HEADERS)
    out: list[TBRow] = []
    for row in rows[header_at + 1 :]:
        code = _cell(row, cols.get("account_code"))
        name = _cell(row, cols.get("account_name"))
        opening = parse_amount(_cell(row, cols.get("opening")))
        debit = parse_amount(_cell(row, cols.get("debit_total")))
        credit = parse_amount(_cell(row, cols.get("credit_total")))
        closing = parse_amount(_cell(row, cols.get("closing")))
        if not code and not name and not (opening or debit or credit or closing):
            continue
        out.append(TBRow(
            account_code=_str_or_none(code),
            account_name=_str_or_none(name),
            opening=opening,
            debit_total=debit,
            credit_total=credit,
            closing=closing,
        ))
    return out


def extract_gl(xlsx_bytes: bytes) -> list[GLRow]:
    rows = _read_first_sheet(xlsx_bytes)
    header_at = _detect_header_row(rows, GL_HEADERS)
    if header_at is None:
        return []
    cols = _header_index([str(c or "") for c in rows[header_at]], GL_HEADERS)
    out: list[GLRow] = []
    for row in rows[header_at + 1 :]:
        debit = parse_amount(_cell(row, cols.get("debit")))
        credit = parse_amount(_cell(row, cols.get("credit")))
        if not debit and not credit:
            amt = parse_amount(_cell(row, cols.get("amount")))
            if amt > 0:
                debit = amt
            elif amt < 0:
                credit = -amt
            else:
                continue
        out.append(GLRow(
            je_number=_str_or_none(_cell(row, cols.get("je_number"))),
            je_date=parse_date(_cell(row, cols.get("je_date"))),
            posting_date=parse_date(_cell(row, cols.get("posting_date"))),
            account_code=_str_or_none(_cell(row, cols.get("account_code"))),
            description=_str_or_none(_cell(row, cols.get("description"))),
            debit=debit,
            credit=credit,
            currency=_str_or_none(_cell(row, cols.get("currency"))),
            preparer=_str_or_none(_cell(row, cols.get("preparer"))),
            approver=_str_or_none(_cell(row, cols.get("approver"))),
        ))
    return out


def _read_first_sheet(xlsx_bytes: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    out: list[list[Any]] = []
    for r in ws.iter_rows(values_only=True):
        out.append(list(r))
    return out


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
