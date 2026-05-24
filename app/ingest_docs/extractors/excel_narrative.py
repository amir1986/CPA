"""Generic XLSX narrative extractor for the USGAAP <> IFRS tool.

Unlike ``excel.extract_trial_balance`` / ``extract_gl`` which produce
typed rows, this one walks every sheet and emits a flat narrative span
per row so financial-statement spreadsheets (which don't follow a fixed
TB/GL shape) still feed the orchestrator.
"""

from __future__ import annotations

import io

from app.ingest_docs.extractors.pdf_text import ExtractedSpan


def extract_xlsx_narrative(body: bytes) -> list[ExtractedSpan]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover — install [parse] extras
        raise RuntimeError("openpyxl not installed (pip install 'cpa[parse]')") from exc

    wb = load_workbook(io.BytesIO(body), data_only=True, read_only=True)
    spans: list[ExtractedSpan] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Take the first non-empty row as a header so per-row spans are
        # self-describing for the LLM. Falls back to col1/col2/... if blank.
        header_row: tuple = next((r for r in rows if any(c is not None and str(c).strip() for c in r)), ())
        header = [str(c).strip() if c is not None else f"col{i + 1}" for i, c in enumerate(header_row)]
        if header:
            spans.append(ExtractedSpan(anchor=f"sheet={sheet_name}.header", text=" | ".join(header)))
        for r_idx, row in enumerate(rows, start=1):
            if row is header_row:
                continue
            cells = [c for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            flat = " | ".join(
                f"{(header[i] if i < len(header) else f'col{i + 1}')}={str(c).strip()}"
                for i, c in enumerate(row)
                if c is not None and str(c).strip()
            )
            if flat:
                spans.append(ExtractedSpan(anchor=f"sheet={sheet_name}.row={r_idx}", text=flat))
    return spans
